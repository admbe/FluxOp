"""Executive workbook export.

One governed .xlsx for leadership: the fiscal-year outlook with a native
Excel chart, budget-group tracking, commitment posture, and the stated
assumptions. Everything in it comes from the same governed reads the
Reports page uses -- the workbook is a presentation of the system of
record, never a separate calculation.
"""

from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

_TITLE_FONT = Font(name="Aptos Display", size=22, bold=True, color="10211C")
_SECTION_FONT = Font(name="Aptos", size=10, bold=True, color="087F5B")
_HEADER_FONT = Font(name="Aptos", bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="12372D")
_WARN_FILL = PatternFill("solid", fgColor="FDE9D9")
_OVER_FILL = PatternFill("solid", fgColor="F8D7DA")
_UNDER_FILL = PatternFill("solid", fgColor="D6F0E0")
_INFO_FILL = PatternFill("solid", fgColor="E8F1FA")
_SUBTLE_FILL = PatternFill("solid", fgColor="F4F7F5")
_MONEY = "$#,##0"
_PERCENT = '0.0"%"'


def _sheet_title(sheet: Worksheet, title: str, subtitle: str) -> int:
    """Use the report's restrained executive hierarchy on every sheet."""
    sheet["A2"] = title
    sheet["A2"].font = _TITLE_FONT
    sheet["A3"] = subtitle
    sheet["A3"].font = Font(name="Aptos", size=10, color="687973")
    sheet.merge_cells("A2:J2")
    sheet.merge_cells("A3:J3")
    sheet.sheet_view.showGridLines = False
    sheet.row_dimensions[2].height = 32
    sheet.row_dimensions[3].height = 20
    return 5


def _header_row(sheet: Worksheet, row: int, headers: list[str]) -> int:
    for index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=row, column=index, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[row].height = 22
    return row + 1


def _autosize(sheet: Worksheet, widths: dict[int, int]) -> None:
    for column, width in widths.items():
        sheet.column_dimensions[get_column_letter(column)].width = width


def _summary_sheet(
    sheet: Worksheet,
    outlook: dict[str, Any],
    commitments: dict[str, Any],
    executive: dict[str, Any],
) -> None:
    currency = outlook.get("currency") or "USD"
    row = _sheet_title(
        sheet,
        f"Azure spend — executive summary ({outlook.get('fiscalYear')})",
        f"Generated {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"
        f" · cost basis {outlook.get('costType')} · currency {currency}"
        " · produced by FluxFinOps governed reporting",
    )
    spend = executive.get("spend") or {}
    anomalies = executive.get("anomalies") or {}
    savings = executive.get("savings") or {}
    coverage = outlook.get("subscriptionCoverage") or {}
    entries = [
        ("Fiscal year", outlook.get("fiscalYear"), None),
        ("FY actual to date", outlook.get("actualToDate"), _MONEY),
        ("FY projected total", outlook.get("fyTotal"), _MONEY),
        (
            "Projection range",
            f"{outlook.get('fyLower'):,.0f} – {outlook.get('fyUpper'):,.0f}",
            None,
        ),
        ("FY budget", outlook.get("fyBudget"), _MONEY),
        ("Variance vs budget", outlook.get("fyVarianceVsBudget"), _MONEY),
        ("Forecast range width", (outlook.get("fyUpper") or 0) - (outlook.get("fyLower") or 0), _MONEY),
        ("Month-to-date spend", spend.get("mtdActual"), _MONEY),
        ("Active cost anomalies", anomalies.get("count"), None),
        ("Realized savings (measured)", savings.get("realizedMonthly"), _MONEY),
        (
            "Active reservations",
            (commitments.get("summary") or {}).get("activeCount"),
            None,
        ),
        (
            "Reservations expiring ≤ 120 days",
            (commitments.get("summary") or {}).get("expiringWithin120Days"),
            None,
        ),
        (
            "Monthly-history coverage",
            f"{coverage.get('covered', 0)} of {coverage.get('configured', 0)}"
            " subscriptions",
            None,
        ),
    ]
    for label, value, number_format in entries:
        sheet.cell(row=row, column=1, value=label).font = Font(bold=True)
        cell = sheet.cell(row=row, column=2, value=value)
        if number_format and isinstance(value, (int, float)):
            cell.number_format = number_format
        if label == "Variance vs budget" and isinstance(value, (int, float)):
            cell.fill = _OVER_FILL if value > 0 else _UNDER_FILL
        if label == "Forecast range width":
            cell.fill = _INFO_FILL
        row += 1
    row += 1
    sheet.cell(row=row, column=1, value="Data limitations").font = _SECTION_FONT
    row += 1
    for limitation in outlook.get("limitations") or ["None recorded."]:
        sheet.cell(row=row, column=1, value=f"• {limitation}")
        row += 1
    uncovered = coverage.get("uncovered") or []
    if uncovered:
        row += 1
        sheet.cell(row=row, column=1, value="Subscriptions requiring cost-history follow-up").font = _SECTION_FONT
        row += 1
        row = _header_row(sheet, row, ["Subscription", "ID", "Last ingestion status", "Message"])
        for item in uncovered:
            sheet.cell(row=row, column=1, value=item.get("label"))
            sheet.cell(row=row, column=2, value=item.get("subscriptionId"))
            sheet.cell(row=row, column=3, value=item.get("lastIngestionStatus") or "No recorded attempt")
            sheet.cell(row=row, column=4, value=item.get("lastIngestionMessage") or "")
            row += 1
    _autosize(sheet, {1: 34, 2: 30, 3: 24, 4: 60})
    sheet.freeze_panes = "A4"


def _outlook_sheet(sheet: Worksheet, outlook: dict[str, Any]) -> None:
    row = _sheet_title(
        sheet,
        f"Fiscal-year outlook — {outlook.get('fiscalYear')}",
        "Actual through the current month; projected thereafter with a lower/"
        "upper planning range. Monthly budget "
        + (
            f"${outlook.get('budgetMonthly'):,.0f}."
            if outlook.get("budgetMonthly") is not None
            else "not configured."
        ),
    )
    headers = ["Month", "Status", "Spend", "Actual", "Lower", "Upper", "Monthly budget", "Variance"]
    data_start = _header_row(sheet, row, headers)
    budget = outlook.get("budgetMonthly")
    current = data_start
    for month in outlook.get("months") or []:
        sheet.cell(row=current, column=1, value=month["month"])
        status = month["status"].replace("inProgress", "in progress")
        sheet.cell(row=current, column=2, value=status)
        is_actual = month["status"] == "actual"
        for column, value in (
            (3, month["amount"]),
            (4, month["amount"] if is_actual else None),
            (5, month["lower"] if not is_actual else None),
            (6, month["upper"] if not is_actual else None),
        ):
            cell = sheet.cell(row=current, column=column, value=value)
            cell.number_format = _MONEY
        if is_actual:
            sheet.cell(row=current, column=2).font = Font(bold=True, color="087F5B")
        elif status == "in progress":
            sheet.cell(row=current, column=2).font = Font(bold=True, color="A56500")
        else:
            sheet.cell(row=current, column=2).font = Font(color="687973")
        if budget is not None:
            cell = sheet.cell(row=current, column=7, value=budget)
            cell.number_format = _MONEY
            if month["amount"] > budget:
                sheet.cell(row=current, column=3).font = Font(bold=True, color="C2410C")
            variance = month["amount"] - budget
            variance_cell = sheet.cell(row=current, column=8, value=variance)
            variance_cell.number_format = _MONEY
            variance_cell.font = Font(
                bold=True,
                color="C2410C" if variance > 0 else "087F5B",
            )
        current += 1

    total_row = current
    sheet.cell(row=total_row, column=1, value="FY TOTAL").font = Font(bold=True)
    for column, value in (
        (3, outlook.get("fyTotal")),
        (7, outlook.get("fyBudget")),
        (8, outlook.get("fyVarianceVsBudget")),
    ):
        cell = sheet.cell(row=total_row, column=column, value=value)
        cell.font = Font(bold=True, color="C2410C" if column == 8 else "10211C")
        cell.number_format = _MONEY
    for column in range(1, 9):
        sheet.cell(row=total_row, column=column).fill = _SUBTLE_FILL

    chart = LineChart()
    chart.title = "Monthly spend vs budget — with planning range"
    chart.height = 12
    chart.width = 24
    chart.y_axis.numFmt = _MONEY
    amounts = Reference(
        sheet, min_col=3, max_col=7, min_row=data_start - 1, max_row=current - 1
    )
    labels = Reference(sheet, min_col=1, min_row=data_start, max_row=current - 1)
    chart.add_data(amounts, titles_from_data=True)
    if len(chart.series) >= 5:
        chart.series[0].tx = SeriesLabel(v="Spend")
        chart.series[0].graphicalProperties.line.solidFill = "188563"
        chart.series[0].graphicalProperties.line.width = 26000
        chart.series[1].tx = SeriesLabel(v="Actual")
        chart.series[1].graphicalProperties.line.solidFill = "188563"
        for index, label in ((2, "Lower range"), (3, "Upper range")):
            chart.series[index].tx = SeriesLabel(v=label)
            chart.series[index].graphicalProperties.line.solidFill = "9DCDBD"
            chart.series[index].graphicalProperties.line.prstDash = "dash"
        chart.series[4].tx = SeriesLabel(v="Monthly budget")
        chart.series[4].graphicalProperties.line.solidFill = "A3ACA8"
        chart.series[4].graphicalProperties.line.prstDash = "dash"
    chart.set_categories(labels)
    sheet.add_chart(chart, f"A{total_row + 2}")
    sheet.freeze_panes = f"A{data_start}"
    _autosize(sheet, {1: 12, 2: 12, 3: 14, 4: 14, 5: 14, 6: 14, 7: 15, 8: 17})


def _groups_sheet(sheet: Worksheet, outlook: dict[str, Any]) -> None:
    row = _sheet_title(
        sheet,
        "Budget groups",
        "Each group tracks its member subscriptions against its own annual"
        " envelope.",
    )
    headers = [
        "Group", "Annual budget", "FY actual to date", "FY projected",
        "Lower", "Upper", "Variance", "Variance %", "Subscriptions backfilled",
    ]
    data_start = _header_row(sheet, row, headers)
    current = data_start
    for group in outlook.get("groups") or []:
        sheet.cell(row=current, column=1, value=group["name"]).font = Font(
            bold=True
        )
        for column, key in (
            (2, "annualBudget"), (3, "actualToDate"), (4, "fyTotal"),
            (5, "fyLower"), (6, "fyUpper"), (7, "variance"),
        ):
            cell = sheet.cell(row=current, column=column, value=group[key])
            cell.number_format = _MONEY
        variance_cell = sheet.cell(row=current, column=7)
        variance_cell.fill = (
            _OVER_FILL if group["variance"] > 0 else _UNDER_FILL
        )
        percent_cell = sheet.cell(
            row=current,
            column=8,
            value=(group["variance"] / group["annualBudget"] if group["annualBudget"] else None),
        )
        percent_cell.number_format = _PERCENT
        sheet.cell(
            row=current,
            column=9,
            value=f"{group['coveredMembers']} of {group['memberCount']}",
        )
        current += 1
    if not (outlook.get("groups") or []):
        sheet.cell(
            row=current,
            column=1,
            value="No budget groups configured yet (Administration →"
            " Budget groups).",
        )
    group_count = len(outlook.get("groups") or [])
    if group_count:
        chart = BarChart()
        chart.type = "bar"
        chart.style = 10
        chart.title = "Annual budget vs FY projected"
        chart.height = 8
        chart.width = 20
        chart.x_axis.numFmt = _MONEY
        labels = Reference(sheet, min_col=1, min_row=data_start, max_row=data_start + group_count - 1)
        for column in (2, 4):
            chart.add_data(Reference(sheet, min_col=column, max_col=column, min_row=data_start - 1, max_row=data_start + group_count - 1), titles_from_data=True)
        if len(chart.series) >= 2:
            chart.series[0].tx = SeriesLabel(v="Annual budget")
            chart.series[1].tx = SeriesLabel(v="FY projected")
            chart.series[0].graphicalProperties.solidFill = "A3ACA8"
            chart.series[1].graphicalProperties.solidFill = "188563"
        chart.set_categories(labels)
        sheet.add_chart(chart, "K4")
    sheet.freeze_panes = f"A{data_start}"
    _autosize(sheet, {1: 18, 2: 15, 3: 17, 4: 14, 5: 14, 6: 14, 7: 14, 8: 12, 9: 24})


def _planning_lens_sheet(
    sheet: Worksheet,
    outlook: dict[str, Any],
    commitments: dict[str, Any],
    executive: dict[str, Any],
) -> None:
    row = _sheet_title(
        sheet,
        "Planning lens — drivers and decisions",
        "A compact view of the signals most likely to change the next planning action.",
    )
    data_start = _header_row(sheet, row, ["Signal", "Value", "Planning context"])
    row = data_start
    spend = executive.get("spend") or {}
    savings = executive.get("savings") or {}
    summary = commitments.get("summary") or {}
    entries = [
        ("FY projected total", outlook.get("fyTotal"), "Governed forecast, including stated assumptions."),
        ("FY variance vs budget", outlook.get("fyVarianceVsBudget"), "Positive values indicate projected overspend."),
        ("Forecast range width", (outlook.get("fyUpper") or 0) - (outlook.get("fyLower") or 0), "Lower width means a tighter planning range."),
        ("Planned savings / month", outlook.get("plannedSavingsMonthly"), "Right-sizing savings included in the outlook."),
        ("Realized savings / month", savings.get("realizedMonthly"), "Measured outcome from implemented actions."),
        ("Active reservations", summary.get("activeCount"), "Review renewal posture alongside utilization."),
        ("MTD actual spend", spend.get("mtdActual"), "Finalized month-to-date spend."),
    ]
    for label, value, context in entries:
        sheet.cell(row=row, column=1, value=label).font = Font(bold=True)
        value_cell = sheet.cell(row=row, column=2, value=value)
        if isinstance(value, (int, float)) and ("reservations" not in label.lower()):
            value_cell.number_format = _MONEY
        if label == "FY variance vs budget" and isinstance(value, (int, float)):
            value_cell.fill = _OVER_FILL if value > 0 else _UNDER_FILL
        sheet.cell(row=row, column=3, value=context).alignment = Alignment(wrap_text=True)
        row += 1
    row += 1
    service_row = _header_row(sheet, row, ["Top MTD service", "MTD actual"])
    services = executive.get("topServices") or []
    for item in services:
        sheet.cell(row=service_row, column=1, value=item["name"])
        cell = sheet.cell(row=service_row, column=2, value=item["mtdActual"])
        cell.number_format = _MONEY
        service_row += 1
    if services:
        chart = BarChart()
        chart.type = "bar"
        chart.title = "Top services — MTD actual"
        chart.height = 8
        chart.width = 20
        chart.x_axis.numFmt = _MONEY
        chart.add_data(Reference(sheet, min_col=2, min_row=service_row - len(services) - 1, max_row=service_row - 1), titles_from_data=True)
        chart.series[0].tx = SeriesLabel(v="MTD actual")
        chart.series[0].graphicalProperties.solidFill = "188563"
        chart.set_categories(Reference(sheet, min_col=1, min_row=service_row - len(services), max_row=service_row - 1))
        sheet.add_chart(chart, "E4")
    sheet.freeze_panes = f"A{data_start}"
    _autosize(sheet, {1: 28, 2: 18, 3: 54})


def _service_composition_sheet(sheet: Worksheet, composition: dict[str, Any]) -> None:
    """Export both billing labels and economic/resource composition."""
    currency = composition.get("currency") or "USD"
    row = _sheet_title(
        sheet,
        "Service composition — billing vs economic",
        f"MTD actual from {composition.get('periodStart', 'current month')} · Currency: {currency}",
    )
    if composition.get("mixedSourceClassification"):
        warning = sheet.cell(row=row, column=1, value="Classification note")
        warning.font = Font(bold=True, color="9C6500")
        warning.fill = _WARN_FILL
        sheet.cell(row=row, column=2, value=composition.get("note", "")).alignment = Alignment(wrap_text=True)
        row += 2

    for title, key in (("Billing-service view", "billingServices"), ("Resource/economic view", "economicCategories"), ("Cost source", "sources")):
        row = _header_row(sheet, row, [title, "MTD actual"])
        for item in composition.get(key) or []:
            sheet.cell(row=row, column=1, value=item.get("name"))
            sheet.cell(row=row, column=2, value=item.get("amount", 0)).number_format = _MONEY
            row += 1
        row += 1
    sheet.cell(row=row, column=1, value="Method note").font = Font(bold=True)
    sheet.cell(row=row, column=2, value=composition.get("note", "")).alignment = Alignment(wrap_text=True)
    sheet.freeze_panes = "A4"
    _autosize(sheet, {1: 32, 2: 20})


def _unit_economics_sheet(sheet: Worksheet, report: dict[str, Any]) -> None:
    summary = report.get("summary") or {}
    row = _sheet_title(
        sheet,
        f"Unit economics — {summary.get('dimensionLabel', 'business dimension')}",
        "Actual month-to-date cost attributed to the configured business dimension.",
    )
    data_start = _header_row(sheet, row, [summary.get("dimensionLabel", "Unit"), "Resources", "Monthly cost", "% of total"])
    row = data_start
    units = report.get("units") or []
    for item in units:
        sheet.cell(row=row, column=1, value=item["name"])
        sheet.cell(row=row, column=2, value=item["resourceCount"])
        sheet.cell(row=row, column=3, value=item["monthlyCost"]).number_format = _MONEY
        sheet.cell(row=row, column=4, value=(item.get("percentOfTotal") or 0) / 100).number_format = _PERCENT
        row += 1
    if units:
        chart = BarChart()
        chart.type = "bar"
        chart.title = "Monthly cost by unit"
        chart.height = 8
        chart.width = 20
        chart.x_axis.numFmt = _MONEY
        chart.add_data(Reference(sheet, min_col=3, min_row=data_start - 1, max_row=row - 1), titles_from_data=True)
        chart.series[0].tx = SeriesLabel(v="Monthly cost")
        chart.series[0].graphicalProperties.solidFill = "2E7D5B"
        chart.set_categories(Reference(sheet, min_col=1, min_row=data_start, max_row=row - 1))
        sheet.add_chart(chart, "F4")
    sheet.cell(row=row + 1, column=1, value="Unattributed cost").font = Font(bold=True)
    sheet.cell(row=row + 1, column=2, value=summary.get("unattributedCost")).number_format = _MONEY
    sheet.freeze_panes = f"A{data_start}"
    _autosize(sheet, {1: 28, 2: 12, 3: 16, 4: 13})


def _commitments_sheet(sheet: Worksheet, commitments: dict[str, Any]) -> None:
    summary = commitments.get("summary") or {}
    row = _sheet_title(
        sheet,
        "Commitments",
        f"{summary.get('activeCount', 0)} active ·"
        f" {summary.get('expiringWithin120Days', 0)} expiring within 120 days"
        f" · fleet 30-day utilization"
        f" {summary.get('averageUtilization30d', '—')}%",
    )
    headers = [
        "Name", "SKU", "Type", "Region", "Qty", "Term", "Scope",
        "Expires", "Days left", "Util 30d %",
    ]
    current = _header_row(sheet, row, headers)
    for item in commitments.get("reservations") or []:
        values = [
            item["name"], item["sku"], item["resourceType"], item["region"],
            item["quantity"], item["term"], item["scopeType"],
            item["expiryDate"], item["daysToExpiry"], item["utilization30d"],
        ]
        for column, value in enumerate(values, start=1):
            sheet.cell(row=current, column=column, value=value)
        days = item["daysToExpiry"]
        if days is not None and days <= 120:
            for column in range(1, len(headers) + 1):
                sheet.cell(row=current, column=column).fill = (
                    _OVER_FILL if days <= 30 else _WARN_FILL
                )
        current += 1
    _autosize(
        sheet,
        {1: 26, 2: 20, 3: 15, 4: 11, 5: 6, 6: 7, 7: 19, 8: 12, 9: 10, 10: 11},
    )


def _assumptions_sheet(sheet: Worksheet, outlook: dict[str, Any]) -> None:
    config = outlook.get("config") or {}
    row = _sheet_title(
        sheet,
        "Assumptions and method",
        "The projection is a planning estimate; these are its recorded"
        " inputs.",
    )
    entries = [
        ("Method", outlook.get("methodVersion")),
        (
            "Seasonal comparison (not primary)",
            (outlook.get("seasonalComparison") or {}).get("fyTotal")
            if outlook.get("seasonalComparison")
            else "—",
        ),
        ("Months of history used", outlook.get("historyMonths")),
        ("Backtest error (MAPE)", outlook.get("backtestMape")),
        (
            "Seasonal comparison YoY factor",
            (outlook.get("seasonalComparison") or {}).get("yoyFactor")
            if outlook.get("seasonalComparison")
            else "—",
        ),
        ("Fiscal year starts", f"Month {config.get('fyStartMonth')}"),
        ("Cost basis", config.get("costType")),
        ("Growth assumption (%/month)", config.get("growthPercentMonthly")),
        (
            "Right-sizing plan savings applied",
            "Yes" if outlook.get("planSavingsApplied") else "No",
        ),
        (
            "Plan savings available ($/month)",
            outlook.get("plannedSavingsMonthly"),
        ),
        ("Savings ramp (months)", config.get("savingsRampMonths")),
        ("Notes", config.get("notes") or "—"),
        ("Assumptions last saved by", config.get("updatedBy") or "—"),
        ("Assumptions last saved at", config.get("updatedAt") or "—"),
    ]
    for label, value in entries:
        sheet.cell(row=row, column=1, value=label).font = Font(bold=True)
        sheet.cell(row=row, column=2, value=value)
        row += 1
    _autosize(sheet, {1: 32, 2: 44})


def build_executive_workbook(database: Any) -> bytes:
    outlook = database.fiscal_year_outlook()
    commitments = database.commitment_inventory()
    executive = database.executive_summary()
    unit_economics = database.unit_economics_report()

    workbook = Workbook()
    _summary_sheet(workbook.active, outlook, commitments, executive)
    workbook.active.title = "Summary"
    _outlook_sheet(workbook.create_sheet("FY outlook"), outlook)
    _groups_sheet(workbook.create_sheet("Budget groups"), outlook)
    _planning_lens_sheet(
        workbook.create_sheet("Planning lens"), outlook, commitments, executive
    )
    if executive.get("serviceComposition"):
        _service_composition_sheet(
            workbook.create_sheet("Service composition"), executive["serviceComposition"]
        )
    _commitments_sheet(workbook.create_sheet("Commitments"), commitments)
    if unit_economics.get("configured"):
        _unit_economics_sheet(workbook.create_sheet("Unit economics"), unit_economics)
    _assumptions_sheet(workbook.create_sheet("Assumptions"), outlook)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
